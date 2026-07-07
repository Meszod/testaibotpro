"""
╔══════════════════════════════════════════════════════════════╗
║      PROFESSIONAL QUIZ BOT — Global Edition (OPTIMIZED)      ║
║   Yaratuvchi: Sultonboyev Muhammad  |  @mesz0d               ║
╚══════════════════════════════════════════════════════════════╝
"""
import asyncio, json, re, os, uuid, logging, io, time
from functools import lru_cache
import aiosqlite, pdfplumber, pytesseract, httpx
from docx import Document
from PIL import Image
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    BufferedInputFile
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError

# ═══════════════════════════════════════════════════════════════
# ⚙️  MUHIT O'ZGARUVCHILARI
# ═══════════════════════════════════════════════════════════════
BOT_TOKEN       = os.getenv("BOT_TOKEN", "")
GROQ_API_KEY    = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL      = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
ADMIN_IDS       = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip()]
ADMIN_USERNAME  = os.getenv("ADMIN_USERNAME", "sltnvv_de")
BOT_USERNAME    = os.getenv("BOT_USERNAME", "teeeestaibot")
DB_FILE         = os.getenv("DB_FILE", "/data/quiz_pro.db")

if not BOT_TOKEN:    raise RuntimeError("BOT_TOKEN topilmadi.")
if not GROQ_API_KEY: raise RuntimeError("GROQ_API_KEY topilmadi.")
if not ADMIN_IDS:    raise RuntimeError("ADMIN_IDS topilmadi.")

os.makedirs(os.path.dirname(DB_FILE) or ".", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("bot.log", encoding="utf-8")]
)
log = logging.getLogger(__name__)

bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
# ⚡ Tezlik uchun: handlerlarni parallel ishlatish
dp  = Dispatcher(storage=MemoryStorage(), handle_in_tasks=True)

# ═══════════════════════════════════════════════════════════════
# 🚀  GLOBAL RESURSLAR (TEZLIK UCHUN)
# ═══════════════════════════════════════════════════════════════
_db: aiosqlite.Connection = None
_http: httpx.AsyncClient  = None

async def get_db() -> aiosqlite.Connection:
    """Global SQLite connection — bir marta ochiladi, tez ishlaydi."""
    global _db
    if _db is None:
        _db = await aiosqlite.connect(DB_FILE)
        await _db.execute("PRAGMA journal_mode=WAL")      # ⚡ Tezlik
        await _db.execute("PRAGMA synchronous=NORMAL")    # ⚡ Tezlik
        await _db.execute("PRAGMA cache_size=10000")      # ⚡ Tezlik
        await _db.execute("PRAGMA temp_store=MEMORY")     # ⚡ Tezlik
    return _db

async def get_http() -> httpx.AsyncClient:
    """Global httpx session — connection reuse."""
    global _http
    if _http is None or _http.is_closed:
        _http = httpx.AsyncClient(timeout=90, limits=httpx.Limits(max_connections=50))
    return _http

async def close_resources():
    global _db, _http
    if _db: await _db.close()
    if _http: await _http.aclose()

# ═══════════════════════════════════════════════════════════════
# 🤖  GROQ API (TEZLASHTIRILGAN)
# ═══════════════════════════════════════════════════════════════
async def groq_chat(prompt: str, max_tokens: int = 4000, temperature: float = 0.1) -> str:
    client = await get_http()
    r = await client.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
        json={
            "model": GROQ_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens, "temperature": temperature
        }
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()

def clean_json(raw: str) -> str:
    raw = re.sub(r"```json|```", "", raw)
    m = re.search(r'\[.*\]', raw, re.DOTALL)
    return m.group(0) if m else raw

# ═══════════════════════════════════════════════════════════════
# 🗄️  DATABASE (OPTIMIZED — global connection)
# ═══════════════════════════════════════════════════════════════
async def init_db():
    db = await get_db()
    await db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        user_id    INTEGER PRIMARY KEY,
        username   TEXT    DEFAULT '',
        first_name TEXT    DEFAULT '',
        last_name  TEXT    DEFAULT '',
        level      TEXT    DEFAULT '',
        correct    INTEGER DEFAULT 0,
        total      INTEGER DEFAULT 0,
        registered INTEGER DEFAULT 0,
        joined_at  TEXT    DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS tests (
        test_id    TEXT PRIMARY KEY,
        owner_id   INTEGER,
        title      TEXT,
        data_json  TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS groups (
        chat_id    INTEGER PRIMARY KEY,
        title      TEXT,
        added_at   TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS results (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        test_id    TEXT,
        user_id    INTEGER,
        score      INTEGER,
        total      INTEGER,
        done_at    TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS referrals (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_id      INTEGER,
        referred_user INTEGER,
        created_at    TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE INDEX IF NOT EXISTS idx_results_test_user ON results(test_id, user_id);
    CREATE INDEX IF NOT EXISTS idx_users_total ON users(total DESC);
    """)
    await db.commit()

# ⚡ Barcha DB funksiyalari global connection ishlatadi
async def upsert_user(uid, username="", first_name="", last_name=""):
    db = await get_db()
    await db.execute("""
        INSERT INTO users (user_id,username,first_name,last_name)
        VALUES(?,?,?,?)
        ON CONFLICT(user_id) DO UPDATE SET
            username=excluded.username,
            first_name=excluded.first_name,
            last_name=excluded.last_name
    """, (uid, username or "", first_name or "", last_name or ""))
    await db.commit()

async def register_user(uid, first_name, last_name, level):
    db = await get_db()
    await db.execute(
        "UPDATE users SET first_name=?,last_name=?,level=?,registered=1 WHERE user_id=?",
        (first_name, last_name, level, uid)
    )
    await db.commit()

async def is_registered(uid) -> bool:
    db = await get_db()
    async with db.execute("SELECT registered FROM users WHERE user_id=?", (uid,)) as cur:
        row = await cur.fetchone()
        return bool(row and row[0])

async def get_user(uid):
    db = await get_db()
    async with db.execute("SELECT * FROM users WHERE user_id=?", (uid,)) as cur:
        return await cur.fetchone()

async def update_stats(uid, correct: int):
    db = await get_db()
    await db.execute(
        "UPDATE users SET correct=correct+?,total=total+1 WHERE user_id=?",
        (correct, uid)
    )
    await db.commit()

async def get_stats(uid):
    db = await get_db()
    async with db.execute("SELECT correct,total FROM users WHERE user_id=?", (uid,)) as cur:
        return await cur.fetchone() or (0, 0)

async def all_users():
    db = await get_db()
    async with db.execute(
        "SELECT user_id,username,first_name,last_name,level,correct,total,joined_at FROM users ORDER BY total DESC"
    ) as cur:
        return await cur.fetchall()

async def save_test(test_id, owner_id, title, data):
    db = await get_db()
    await db.execute(
        "INSERT OR REPLACE INTO tests(test_id,owner_id,title,data_json) VALUES(?,?,?,?)",
        (test_id, owner_id, title, json.dumps(data, ensure_ascii=False))
    )
    await db.commit()

async def get_test(test_id):
    db = await get_db()
    async with db.execute(
        "SELECT title,data_json,owner_id FROM tests WHERE test_id=?", (test_id,)
    ) as cur:
        row = await cur.fetchone()
        if row:
            return row[0], json.loads(row[1]), row[2]
    return None, None, None

async def save_group(chat_id, title):
    db = await get_db()
    await db.execute("INSERT OR IGNORE INTO groups(chat_id,title) VALUES(?,?)", (chat_id, title))
    await db.commit()

async def get_groups():
    db = await get_db()
    async with db.execute("SELECT chat_id,title FROM groups") as cur:
        return await cur.fetchall()

async def save_result(test_id, user_id, score, total):
    db = await get_db()
    await db.execute(
        "INSERT INTO results(test_id,user_id,score,total) VALUES(?,?,?,?)",
        (test_id, user_id, score, total)
    )
    await db.commit()

async def get_user_result(test_id, user_id):
    db = await get_db()
    async with db.execute(
        "SELECT score,total,done_at FROM results WHERE test_id=? AND user_id=? ORDER BY id ASC LIMIT 1",
        (test_id, user_id)
    ) as cur:
        return await cur.fetchone()

async def get_leaderboard(limit: int = 10):
    db = await get_db()
    async with db.execute(
        """SELECT user_id,username,first_name,last_name,correct,total
           FROM users WHERE total>0
           ORDER BY correct DESC, total ASC LIMIT ?""",
        (limit,)
    ) as cur:
        return await cur.fetchall()

async def get_all_tests():
    db = await get_db()
    async with db.execute(
        "SELECT test_id,title,data_json,created_at FROM tests ORDER BY created_at DESC"
    ) as cur:
        return await cur.fetchall()

async def delete_test(test_id):
    db = await get_db()
    await db.execute("DELETE FROM tests WHERE test_id=?", (test_id,))
    await db.execute("DELETE FROM results WHERE test_id=?", (test_id,))
    await db.commit()

# ═══════════════════════════════════════════════════════════════
# 🔗  REFERAL
# ═══════════════════════════════════════════════════════════════
async def get_referral_count(uid: int) -> int:
    db = await get_db()
    async with db.execute("SELECT COUNT(*) FROM referrals WHERE owner_id=?", (uid,)) as cur:
        result = await cur.fetchone()
        return result[0] if result else 0

async def add_referral(owner_id: int, referred_user: int):
    db = await get_db()
    async with db.execute(
        "SELECT id FROM referrals WHERE owner_id=? AND referred_user=?",
        (owner_id, referred_user)
    ) as cur:
        if not await cur.fetchone():
            await db.execute(
                "INSERT INTO referrals(owner_id, referred_user) VALUES(?, ?)",
                (owner_id, referred_user)
            )
            await db.commit()
            log.info(f"Referral: {owner_id} -> {referred_user}")

# ═══════════════════════════════════════════════════════════════
# 📢  ADMIN NOTIFICATION (⚡ ORQADA ISHLAYDI)
# ═══════════════════════════════════════════════════════════════
async def notify_admins(uid: int, test_title: str, score: int, total: int):
    """⚡ Bu funksiya endi asosiy oqimni to'smaydi — asyncio.create_task orqali."""
    pct = round(score / total * 100) if total else 0
    medal = ("🏆 Mukammal!" if pct == 100 else
             "🥇 A'lo!" if pct >= 80 else
             "🥈 Yaxshi!" if pct >= 60 else
             "🥉 O'rtacha" if pct >= 40 else
             "📚 Past natija")

    u = await get_user(uid)
    username = f"@{u[1]}" if u and u[1] else None
    full_name = f"{u[2] or ''} {u[3] or ''}".strip() if u else "-"
    level = u[4] if u and u[4] else "-"
    correct, total_all = await get_stats(uid)

    text = (
        f"📬 <b>Yangi test natijasi!</b>\n\n"
        f"👤 <b>O'quvchi:</b> {full_name}\n"
        f"📱 Username: {username or '-'}\n"
        f"🆔 ID: <code>{uid}</code>\n"
        f"📚 Daraja: {level}\n\n"
        f"📝 <b>Test:</b> {test_title}\n\n"
        f"📊 <b>Natija:</b>\n"
        f"✅ To'g'ri: <b>{score}</b>\n"
        f"❌ Noto'g'ri: <b>{total - score}</b>\n"
        f"📝 Jami savollar: <b>{total}</b>\n"
        f"🎯 Foiz: <b>{pct}%</b>\n"
        f"{medal}\n\n"
        f"📈 <b>Umumiy statistika:</b>\n"
        f"✅ To'g'ri: {correct} | 📝 Jami: {total_all}\n\n"
        f"🕒 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )

    # ✅ Tuzatilgan markup
    kb = None
    if username:
        kb = InlineKeyboardBuilder()
        kb.button(text="👤 Profilni ko'rish", url=f"https://t.me/{username[1:]}")

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text, reply_markup=kb.as_markup() if kb else None)
        except Exception as e:
            log.warning(f"Admin {admin_id} notification error: {e}")

# ═══════════════════════════════════════════════════════════════
# 📊  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════
def build_excel(users: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistika"
    headers = ["#", "ID", "Username", "Ism", "Familya", "Daraja",
               "To'g'ri", "Jami", "Foiz(%)", "Qo'shilgan"]
    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for i, u in enumerate(users, 1):
        uid, uname, fname, lname, level, correct, total, joined = u
        pct = round(correct / total * 100) if total else 0
        row = [i, uid, f"@{uname}" if uname else "-",
               fname, lname, level, correct, total, pct, joined]
        for col, val in enumerate(row, 1):
            ws.cell(row=i+1, column=col, value=val)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_txt(users: list) -> str:
    lines = [f"📊 Statistika — {datetime.now().strftime('%d.%m.%Y %H:%M')}\n",
             f"Jami foydalanuvchilar: {len(users)}\n",
             "─" * 55]
    for i, u in enumerate(users, 1):
        uid, uname, fname, lname, level, correct, total, joined = u
        pct = round(correct / total * 100) if total else 0
        lines.append(
            f"{i}. {fname} {lname} (@{uname or '-'})\n"
            f"   Daraja: {level or '-'}  |  {correct}/{total} ({pct}%)\n"
            f"   ID: {uid}  |  {joined}"
        )
    return "\n".join(lines)

# ═══════════════════════════════════════════════════════════════
# 📂  FAYL O'QISH (⚡ PARALLEL — asyncio.to_thread)
# ═══════════════════════════════════════════════════════════════
ALLOWED_EXT = {".pdf",".docx",".txt",".md",".jpg",".jpeg",".png",".bmp",".tiff",".webp"}

def _extract_pdf(path: str) -> str:
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if not t or len(t.strip()) < 20:
                try:
                    img = page.to_image(resolution=250).original
                    t = pytesseract.image_to_string(img, lang="uzb+rus+eng")
                except Exception as e:
                    log.warning(f"OCR error: {e}")
                    t = ""
            text += (t or "") + "\n"
    return text

def _extract_docx(path: str) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)

def _extract_text_file(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def _extract_image(path: str) -> str:
    img = Image.open(path)
    return pytesseract.image_to_string(img, lang="uzb+rus+eng")

async def extract_text(path: str) -> str:
    """⚡ Blocking funksiyalarni thread ichida ishlatadi — bot qotmaydi."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".pdf":
            return await asyncio.to_thread(_extract_pdf, path)
        elif ext == ".docx":
            return await asyncio.to_thread(_extract_docx, path)
        elif ext in (".txt", ".md"):
            return await asyncio.to_thread(_extract_text_file, path)
        elif ext in (".jpg",".jpeg",".png",".bmp",".tiff",".webp"):
            return await asyncio.to_thread(_extract_image, path)
    except Exception as e:
        log.warning(f"extract_text({ext}): {e}")
    return ""

# ═══════════════════════════════════════════════════════════════
# 🔍  SAVOL PARSER
# ═══════════════════════════════════════════════════════════════
def parse_questions(text: str) -> list:
    questions = []
    blocks = re.split(r'\n(?=\d{1,3}[\.\)]\s)', text)
    for b in blocks:
        q = re.search(r'\d{1,3}[\.\)]\s*(.+)', b)
        if not q:
            continue
        opts = re.findall(r'(?:^|\n)\s*([A-Da-d][\.\)]\s*.+)', b, re.MULTILINE)
        opts = [re.sub(r'^[A-Da-d][\.\)]\s*', '', o).strip() for o in opts]
        cm = re.search(r"(?:tog.ri|javob|answer|correct)[:\s]*([A-Da-d])", b, re.IGNORECASE)
        cidx = "abcd".index(cm.group(1).lower()) if cm else 0
        if len(opts) >= 2:
            questions.append({
                "question": q.group(1).strip(),
                "options": opts[:4],
                "correct": cidx
            })
    return questions

async def ai_fix(data: list) -> list:
    prompt = f"""Test savollarini tekshir va to'g'irla.
Har bir savolda: question(string), options(4ta), correct(0-3 int).
Faqat JSON array qaytar:
{json.dumps(data, ensure_ascii=False, indent=2)}"""
    try:
        raw = await groq_chat(prompt, 4000, 0.1)
        fixed = json.loads(clean_json(raw))
        if isinstance(fixed, list) and fixed:
            return fixed
    except Exception as e:
        log.warning(f"ai_fix: {e}")
    return data

async def ai_extract(text: str) -> list:
    prompt = f"""Quyidagi matndagi test savollarini topib JSON formatda chiqar.
Format: [{{"question":"...","options":["A","B","C","D"],"correct":0}}]
Faqat JSON:
{text[:3000]}"""
    try:
        raw = await groq_chat(prompt, 3000, 0.1)
        return json.loads(clean_json(raw))
    except Exception as e:
        log.warning(f"ai_extract: {e}")
    return []

async def ai_generate(topic: str, count: int = 10) -> list:
    prompt = f""""{topic}" mavzusida {count} ta test savolini yaratgin.
Format: [{{"question":"...","options":["A","B","C","D"],"correct":0}}]
Faqat JSON:"""
    try:
        raw = await groq_chat(prompt, 4000, 0.7)
        return json.loads(clean_json(raw))
    except Exception as e:
        log.warning(f"ai_generate: {e}")
    return []

# ═══════════════════════════════════════════════════════════════
# 🎮  SESSION
# ═══════════════════════════════════════════════════════════════
sessions: dict = {}

# ═══════════════════════════════════════════════════════════════
# ⌨️  KLAVIATURALAR
# ═══════════════════════════════════════════════════════════════
LETTERS = ["A", "B", "C", "D"]

def answer_kb(options: list) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for i, opt in enumerate(options):
        b.button(text=f"{LETTERS[i]}) {opt[:55]}", callback_data=f"ans_{i}")
    b.adjust(1)
    return b.as_markup()

def admin_menu_kb() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📄 Fayl yuklash", callback_data="mode_file")
    b.button(text="✍️ Matn kiritish", callback_data="mode_text")
    b.button(text="🤖 AI bilan test yaratish", callback_data="mode_ai")
    b.button(text="📊 Mening statistikam", callback_data="my_stats")
    b.button(text="🏆 Reyting", callback_data="leaderboard")
    b.button(text="🔗 Taklif qilish", callback_data="my_referral")
    b.button(text="👨‍💼 Admin panel", callback_data="admin_panel")
    b.adjust(2)
    return b.as_markup()

def user_menu_kb() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📝 Testni boshlash", callback_data="start_user_test")
    b.button(text="📊 Mening statistikam", callback_data="my_stats")
    b.button(text="🏆 Reyting", callback_data="leaderboard")
    b.button(text="👥 Taklif qilish", callback_data="my_referral")
    b.adjust(2)
    return b.as_markup()

def home_kb() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(1)
    return b.as_markup()

def level_kb() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for lvl in ["🟢 Boshlang'ich (A1-A2)", "🟡 O'rta (B1-B2)", "🔴 Yuqori (C1-C2)"]:
        b.button(text=lvl, callback_data=f"lvl_{lvl}")
    b.adjust(1)
    return b.as_markup()

def share_test_kb(test_id: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📤 Guruhga tashlash", callback_data=f"sendtogroup_{test_id}")
    b.button(text="🔗 Linkni nusxalash", callback_data=f"copylink_{test_id}")
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(2)
    return b.as_markup()

def group_list_kb(groups: list, test_id: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for chat_id, title in groups:
        b.button(text=f"💬 {title}", callback_data=f"tosend_{chat_id}_{test_id}")
    b.button(text="🔙 Orqaga", callback_data=f"back_share_{test_id}")
    b.adjust(1)
    return b.as_markup()

def group_start_kb(test_id: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📝 Testni yechish", url=f"https://t.me/{BOT_USERNAME}?start=test_{test_id}")
    b.adjust(1)
    return b.as_markup()

def admin_kb() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📊 Statistika (Excel)", callback_data="admin_excel")
    b.button(text="📝 Statistika (Text)", callback_data="admin_txt")
    b.button(text="👥 Foydalanuvchilar soni", callback_data="admin_count")
    b.button(text="🏆 Reyting", callback_data="leaderboard")
    b.button(text="📋 Testlar ro'yxati", callback_data="admin_tests")
    b.button(text="🔍 Foydalanuvchi qidirish", callback_data="admin_search_user")
    b.button(text="📢 Hammaga xabar", callback_data="admin_broadcast")
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(2)
    return b.as_markup()

def already_done_kb(test_id: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="🔁 Tayyorlanish uchun yechish", callback_data=f"practice_{test_id}")
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(1)
    return b.as_markup()

# ═══════════════════════════════════════════════════════════════
# 🎮  QUIZ CORE
# ═══════════════════════════════════════════════════════════════
async def start_quiz(chat_id: int, uid: int, tests: list,
                     title: str = "Test", test_id: str = None, practice: bool = False):
    sessions[uid] = {
        "tests": tests, "index": 0, "score": 0,
        "title": title, "test_id": test_id,
        "practice": practice, "started_at": time.time()
    }
    await send_question(chat_id, uid)

async def send_already_done_message(target, uid: int, test_id: str, title: str):
    result = await get_user_result(test_id, uid)
    score, total, done_at = result if result else (0, 0, "-")
    pct = round(score / total * 100) if total else 0
    text = (
        f"⚠️ <b>Siz bu testni allaqachon yechgansiz!</b>\n\n"
        f"📝 <b>{title}</b>\n\n"
        f"📊 <b>Sizning natijangiz:</b>\n"
        f"✅ To'g'ri: <b>{score}</b>\n"
        f"❌ Noto'g'ri: <b>{total - score}</b>\n"
        f"🎯 Foiz: <b>{pct}%</b>\n"
        f"🕒 Sana: {done_at}\n\n"
        f"ℹ️ Har bir test faqat <b>1 marta</b> ball uchun hisoblanadi.\n"
        f"Agar mashq qilib ko'rmoqchi bo'lsangiz, quyidagi tugmani bosing —\n"
        f"bu safar javoblaringiz <b>reytingga hisoblanmaydi</b>."
    )
    await target.answer(text, reply_markup=already_done_kb(test_id))

async def send_question(chat_id: int, uid: int):
    d = sessions[uid]
    idx = d["index"]
    total = len(d["tests"])
    q = d["tests"][idx]
    await bot.send_message(
        chat_id,
        f"📌 <b>Savol {idx+1}/{total}</b>\n\n{q['question']}",
        reply_markup=answer_kb(q["options"])
    )

async def process_text(message: Message, text: str, title: str = "Test"):
    uid = message.from_user.id
    status = await message.answer("🔍 Savollar aniqlanmoqda...")
    parsed = parse_questions(text)
    if not parsed:
        await status.edit_text("🤖 AI yordamida izlanmoqda...")
        parsed = await ai_extract(text)
    if not parsed:
        await status.edit_text(
            "❌ Savollar topilmadi.\n\n"
            "Format:\n<code>1. Savol?\nA) ...\nB) ...\nC) ...\nD) ...</code>"
        )
        return
    await status.edit_text(f"✅ {len(parsed)} ta savol topildi.\n🧠 AI tekshirmoqda...")
    tests = await ai_fix(parsed)
    test_id = str(uuid.uuid4())[:8]
    await save_test(test_id, uid, title, tests)
    await status.edit_text(
        f"🎉 <b>Test tayyor!</b>\n📝 {len(tests)} ta savol\n🏷 {title}",
        reply_markup=share_test_kb(test_id)
    )

# ══════════════════════════════════════════════════════════════
# 📌  FSM
# ═══════════════════════════════════════════════════════════════
class S(StatesGroup):
    wait_file     = State()
    wait_text     = State()
    wait_ai_topic = State()
    reg_first     = State()
    reg_last      = State()
    reg_level     = State()
    broadcast     = State()

# ═══════════════════════════════════════════════════════════════
# 🚦  /start
# ═══════════════════════════════════════════════════════════════
@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    uid = message.from_user.id
    args = message.text.split(maxsplit=1)[1] if " " in message.text else ""

    # ⚡ Faqat yangi foydalanuvchilarni saqlash
    if not await is_registered(uid):
        await upsert_user(uid, message.from_user.username,
                          message.from_user.first_name, message.from_user.last_name)

    if message.chat.type in ("group", "supergroup"):
        await save_group(message.chat.id, message.chat.title or "Guruh")

    if args.startswith("test_"):
        test_id = args[5:]
        title, tests, owner = await get_test(test_id)
        if not tests:
            return await message.answer("❌ Test topilmadi yoki o'chirilgan.")
        if not await is_registered(uid):
            await state.set_state(S.reg_first)
            await state.update_data(pending_test=test_id)
            return await message.answer(
                "👋 Xush kelibsiz!\n\n"
                "Testni boshlashdan oldin ro'yxatdan o'ting.\n\n"
                "📝 <b>Ismingizni kiriting:</b>"
            )
        existing = await get_user_result(test_id, uid)
        if existing:
            return await send_already_done_message(message, uid, test_id, title)

        u = await get_user(uid)
        await message.answer(
            f"🎯 <b>{title}</b>\n"
            f"📝 {len(tests)} ta savol\n\n"
            f"Salom, <b>{u[2]} {u[3]}</b>!\n"
            f"Test boshlanmoqda..."
        )
        await start_quiz(message.chat.id, uid, tests, title, test_id)
        return

    if args.startswith("ref_"):
        ref_code = args[4:]
        try:
            owner_id = int(ref_code)
            if owner_id != uid:
                await add_referral(owner_id, uid)
        except ValueError:
            log.warning(f"Noto'g'ri referral code: {ref_code}")

    if not await is_registered(uid):
        await state.set_state(S.reg_first)
        return await message.answer(
            "👋 <b>Xush kelibsiz!</b>\n\n"
            "Botdan foydalanish uchun qisqacha ro'yxatdan o'ting.\n\n"
            "📝 <b>Ismingizni kiriting:</b>"
        )

    u = await get_user(uid)
    is_admin_user = uid in ADMIN_IDS
    menu_kb = admin_menu_kb() if is_admin_user else user_menu_kb()

    await message.answer(
        f"👋 Salom, <b>{u[2]} {u[3]}</b>!\n\n"
        f"🤖 <b>Quiz Bot</b> — professional test platformasi.\n\n"
        f"Nima qilmoqchisiz? 👇",
        reply_markup=menu_kb
    )

# ═══════════════════════════════════════════════════════════════
# 📝  RO'YXATDAN O'TISH
# ═══════════════════════════════════════════════════════════════
@dp.message(S.reg_first, F.text)
async def reg_first(message: Message, state: FSMContext):
    await state.update_data(first_name=message.text.strip())
    await state.set_state(S.reg_last)
    await message.answer("✅ Yaxshi!\n\n📝 <b>Familyangizni kiriting:</b>")

@dp.message(S.reg_last, F.text)
async def reg_last(message: Message, state: FSMContext):
    await state.update_data(last_name=message.text.strip())
    await state.set_state(S.reg_level)
    await message.answer(
        "✅ Zo'r!\n\n📊 <b>Nemis tili bilim darajangizni tanlang:</b>",
        reply_markup=level_kb()
    )

@dp.callback_query(S.reg_level, F.data.startswith("lvl_"))
async def reg_level(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass

    level = c.data[4:]
    data = await state.get_data()
    await state.clear()
    uid = c.from_user.id
    await register_user(uid, data["first_name"], data["last_name"], level)
    pending = data.get("pending_test")

    try:
        await c.message.edit_text(
            f"🎉 <b>Ro'yxatdan o'tdingiz!</b>\n\n"
            f"👤 {data['first_name']} {data['last_name']}\n"
            f"📊 Daraja: {level}"
        )
    except TelegramBadRequest:
        pass

    if pending:
        title, tests, owner = await get_test(pending)
        if tests:
            existing = await get_user_result(pending, uid)
            if existing:
                await send_already_done_message(c.message, uid, pending, title)
                return
            await c.message.answer(
                f"🎯 <b>{title}</b> — {len(tests)} ta savol\n\nTest boshlanmoqda..."
            )
            await start_quiz(c.message.chat.id, uid, tests, title, pending)
            return

    is_admin_user = uid in ADMIN_IDS
    menu_kb = admin_menu_kb() if is_admin_user else user_menu_kb()

    await c.message.answer(
        "Endi botdan to'liq foydalanishingiz mumkin! 🚀",
        reply_markup=menu_kb
    )

# ═══════════════════════════════════════════════════════════════
# 🏠  ASOSIY MENYU
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data == "home")
async def cb_home(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass

    await state.clear()
    u = await get_user(c.from_user.id)
    name = f"{u[2]} {u[3]}" if u and len(u) > 2 else c.from_user.first_name
    is_admin_user = c.from_user.id in ADMIN_IDS
    menu_kb = admin_menu_kb() if is_admin_user else user_menu_kb()

    try:
        await c.message.edit_text(
            f"👋 Salom, <b>{name}</b>!\n\nNima qilmoqchisiz? 👇",
            reply_markup=menu_kb
        )
    except TelegramBadRequest:
        await c.message.answer(
            f"👋 Salom, <b>{name}</b>!\n\nNima qilmoqchisiz? 👇",
            reply_markup=menu_kb
        )

# ═══════════════════════════════════════════════════════════════
# 📁  REJIMLAR
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data == "mode_file")
async def cb_file(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass
    if c.from_user.id not in ADMIN_IDS:
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")
    await state.set_state(S.wait_file)
    text = "📎 <b>Faylni yuboring:</b>\n\n• PDF (skanerlangan ham)\n• DOCX  • TXT  • Rasm (JPG/PNG)\n\nFaylni shu chatga tashlang 👇"
    try: await c.message.edit_text(text)
    except TelegramBadRequest: await c.message.answer(text)

@dp.callback_query(F.data == "mode_text")
async def cb_text(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass
    if c.from_user.id not in ADMIN_IDS:
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")
    await state.set_state(S.wait_text)
    text = (
        "✍️ <b>Savollarni matn ko'rinishida yuboring.</b>\n\n"
        "Namuna:\n"
        "<code>1. Python nima?\n"
        "A) Dasturlash tili\nB) OS\nC) MB\nD) Protokol</code>"
    )
    try: await c.message.edit_text(text)
    except TelegramBadRequest: await c.message.answer(text)

@dp.callback_query(F.data == "mode_ai")
async def cb_ai(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass
    if c.from_user.id not in ADMIN_IDS:
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")
    await state.set_state(S.wait_ai_topic)
    text = (
        "🤖 <b>Qaysi mavzuda test yaratay?</b>\n\n"
        "Misol: <i>Python dasturlash</i>, <i>O'zbekiston tarixi</i>\n\n"
        "Mavzuni yozing 👇"
    )
    try: await c.message.edit_text(text)
    except TelegramBadRequest: await c.message.answer(text)

@dp.callback_query(F.data == "start_user_test")
async def cb_start_user_test(c: CallbackQuery):
    try: await c.answer()
    except: pass

    db = await get_db()
    async with db.execute(
        "SELECT test_id, title, data_json FROM tests ORDER BY created_at DESC LIMIT 10"
    ) as cur:
        tests = await cur.fetchall()

    if not tests:
        return await c.message.answer(
            "📝 Hozircha testlar mavjud emas.\n"
            "Admin test yaratgandan keyin bu yerda ko'rinadi.",
            reply_markup=home_kb()
        )

    b = InlineKeyboardBuilder()
    for test_id, title, data_json in tests:
        try:
            data = json.loads(data_json)
            count = len(data)
        except:
            count = "?"
        b.button(text=f"📝 {title[:40]} ({count} savol)", callback_data=f"play_{test_id}")
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(1)

    try:
        await c.message.edit_text(
            "📝 <b>Mavjud testlar:</b>\n\nTestni tanlang 👇",
            reply_markup=b.as_markup()
        )
    except TelegramBadRequest:
        await c.message.answer(
            "📝 <b>Mavjud testlar:</b>\n\nTestni tanlang 👇",
            reply_markup=b.as_markup()
        )

@dp.callback_query(F.data.startswith("play_"))
async def cb_play_test(c: CallbackQuery):
    try: await c.answer()
    except: pass

    test_id = c.data.split("play_")[1]
    title, tests, owner = await get_test(test_id)
    if not tests:
        return await c.message.answer("❌ Test topilmadi.", reply_markup=home_kb())

    uid = c.from_user.id
    existing = await get_user_result(test_id, uid)
    if existing:
        return await send_already_done_message(c.message, uid, test_id, title)

    await c.message.answer(
        f"🎯 <b>{title}</b>\n"
        f"📝 {len(tests)} ta savol\n\n"
        f"Test boshlanmoqda..."
    )
    await start_quiz(c.message.chat.id, uid, tests, title, test_id)

# ═══════════════════════════════════════════════════════════════
# 🔁  TAYYORLANISH REJIMI
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data.startswith("practice_"))
async def cb_practice_test(c: CallbackQuery):
    try: await c.answer()
    except: pass

    test_id = c.data.split("practice_")[1]
    title, tests, owner = await get_test(test_id)
    if not tests:
        return await c.message.answer("❌ Test topilmadi.", reply_markup=home_kb())

    await c.message.answer(
        f"🔁 <b>Tayyorlanish rejimi</b>\n\n"
        f"🎯 <b>{title}</b>\n"
        f"📝 {len(tests)} ta savol\n\n"
        f"⚠️ Diqqat: bu urinishning ballari <b>reytingga hisoblanmaydi</b>, "
        f"faqat mashq qilish uchun.\n\nTest boshlanmoqda..."
    )
    await start_quiz(c.message.chat.id, c.from_user.id, tests, title, test_id, practice=True)

# ═══════════════════════════════════════════════════════════════
# 🔗  REFERAL
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data == "my_referral")
async def cb_my_referral(c: CallbackQuery):
    try: await c.answer()
    except: pass

    uid = c.from_user.id
    ref_link = f"https://t.me/{BOT_USERNAME}?start=ref_{uid}"
    referral_count = await get_referral_count(uid)

    b = InlineKeyboardBuilder()
    b.button(text="📤 Linkni nusxalash", switch_inline_query=f"🎯 Quiz Botga qo'shiling!\n\n{ref_link}")
    b.button(text="🏠 Asosiy menyu", callback_data="home")
    b.adjust(1)

    text = (
        f"🔗 <b>Sizning referral linkingiz:</b>\n\n"
        f"<code>{ref_link}</code>\n\n"
        f"📊 <b>Statistika:</b>\n"
        f"👥 Taklif qilinganlar: <b>{referral_count} ta</b>\n\n"
        f"Bu linkni do'stlaringizga yuboring!\n"
        f"Ular botga kirganda sizning hisobingizga qo'shiladi."
    )
    try: await c.message.edit_text(text, reply_markup=b.as_markup())
    except TelegramBadRequest: await c.message.answer(text, reply_markup=b.as_markup())

# ═══════════════════════════════════════════════════════════════
# 📂  FAYL HANDLER
# ═══════════════════════════════════════════════════════════════
@dp.message(S.wait_file, F.document | F.photo)
async def handle_file(message: Message, state: FSMContext):
    await state.clear()
    if message.photo:
        fobj, fname = message.photo[-1], "photo.jpg"
    else:
        fobj, fname = message.document, (message.document.file_name or "file.txt").lower()
    ext = os.path.splitext(fname)[1].lower()
    if ext not in ALLOWED_EXT:
        return await message.answer(
            f"❌ Format qo'llab-quvvatlanmaydi.\n"
            f"Qabul qilinadi: {', '.join(sorted(ALLOWED_EXT))}"
        )
    
    # ⚡ Fayl hajmini tekshirish (20 MB limit)
    if hasattr(fobj, 'file_size') and fobj.file_size and fobj.file_size > 20 * 1024 * 1024:
        return await message.answer("❌ Fayl hajmi 20 MB dan oshmasligi kerak!")

    try:
        tg_file = await bot.get_file(fobj.file_id)
        path = f"/tmp/qb_{message.from_user.id}{ext}"
        await bot.download_file(tg_file.file_path, path)
    except Exception as e:
        log.error(f"Fayl yuklashda xato: {e}")
        return await message.answer("❌ Faylni yuklab olishda xato. Qayta urining.")

    await message.answer("📖 Fayl o'qilmoqda...")
    text = await extract_text(path)
    try: os.remove(path)
    except: pass
    if not text:
        return await message.answer("❌ Fayldan matn o'qilmadi. Boshqa format sinab ko'ring.")
    await process_text(message, text, title=fname)

@dp.message(S.wait_text, F.text)
async def handle_text(message: Message, state: FSMContext):
    await state.clear()
    await process_text(message, message.text, title="Matn testi")

@dp.message(S.wait_ai_topic, F.text)
async def handle_ai(message: Message, state: FSMContext):
    await state.clear()
    topic = message.text.strip()
    status = await message.answer(f"🤖 <b>{topic}</b> mavzusida test yaratilmoqda...")
    tests = await ai_generate(topic, count=10)
    if not tests:
        return await status.edit_text("❌ AI test yarata olmadi. Qayta urining.")
    test_id = str(uuid.uuid4())[:8]
    await save_test(test_id, message.from_user.id, topic, tests)
    await status.edit_text(
        f"🎉 <b>Test tayyor!</b>\n📝 {len(tests)} ta savol\n🏷 {topic}",
        reply_markup=share_test_kb(test_id)
    )

# ═══════════════════════════════════════════════════════════════
# 🔗  TEST ULASHISH
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data.startswith("copylink_"))
async def cb_copy_link(c: CallbackQuery):
    try: await c.answer()
    except: pass
    test_id = c.data.split("copylink_")[1]
    link = f"https://t.me/{BOT_USERNAME}?start=test_{test_id}"
    try:
        await c.message.answer(f"🔗 <b>Test linki:</b>\n\n<code>{link}</code>\n\nBu linkni do'stlaringizga yuboring!")
    except: pass

@dp.callback_query(F.data.startswith("sendtogroup_"))
async def cb_send_to_group(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if c.from_user.id not in ADMIN_IDS:
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")

    test_id = c.data.split("sendtogroup_")[1]
    groups = await get_groups()
    if not groups:
        return await c.message.answer(
            "❗ Hali hech qanday guruh yo'q.\n"
            "Botni guruhga qo'shing va admin qiling!"
        )

    text = "💬 <b>Qaysi guruhga test yubormoqchisiz?</b>\n\nPastdan tanlang 👇"
    try: await c.message.edit_text(text, reply_markup=group_list_kb(groups, test_id))
    except TelegramBadRequest: await c.message.answer(text, reply_markup=group_list_kb(groups, test_id))

@dp.callback_query(F.data.startswith("back_share_"))
async def cb_back_share(c: CallbackQuery):
    try: await c.answer()
    except: pass
    test_id = c.data.split("back_share_")[1]
    try: await c.message.edit_text("Test ulashish:", reply_markup=share_test_kb(test_id))
    except TelegramBadRequest: await c.message.answer("Test ulashish:", reply_markup=share_test_kb(test_id))

@dp.callback_query(F.data.startswith("tosend_"))
async def cb_to_send(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if c.from_user.id not in ADMIN_IDS:
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")

    parts = c.data.split("_")
    chat_id = int(parts[1])
    test_id = parts[2]

    try:
        bot_member = await bot.get_chat_member(chat_id, bot.id)
        if bot_member.status not in ["administrator", "creator"]:
            return await c.message.answer(
                "❗ Bot guruhda admin emas!\n"
                "Botga admin huquqi bering va qayta urining."
            )
    except TelegramForbiddenError:
        return await c.message.answer("❗ Bot guruhdan chiqarib yuborilgan!")
    except Exception as e:
        log.warning(f"admin check error: {e}")
        return await c.message.answer(f"❗ Xato: {str(e)[:100]}")

    title, tests, owner = await get_test(test_id)
    if not tests:
        return await c.message.answer("❌ Test topilmadi.")

    bot_link = f"https://t.me/{BOT_USERNAME}?start=test_{test_id}"

    try:
        await bot.send_message(
            chat_id,
            f"📢 <b>Yangi test e'lon qilindi!</b>\n\n"
            f"📝 <b>{title}</b>\n"
            f"🔢 Savollar soni: <b>{len(tests)} ta</b>\n\n"
            f"✅ Testni yechish uchun quyidagi tugmani bosing!\n"
            f"(Bot sizdan ism-familya so'raydi, keyin test boshlanadi)\n\n"
            f"🔗 Yoki link: <code>{bot_link}</code>",
            reply_markup=group_start_kb(test_id)
        )

        text = f"✅ <b>Test muvaffaqiyatli yuborildi!</b>\n\n💬 Guruh: {chat_id}\n📝 {len(tests)} ta savol"
        try: await c.message.edit_text(text, reply_markup=home_kb())
        except TelegramBadRequest: await c.message.answer(text, reply_markup=home_kb())
    except TelegramForbiddenError:
        await c.message.answer("❗ Bot guruhdan chiqarib yuborilgan!")
    except Exception as e:
        await c.message.answer(f"Xato: {str(e)[:50]}")

# ═══════════════════════════════════════════════════════════════
# 📊  STATISTIKA
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data == "my_stats")
async def cb_my_stats(c: CallbackQuery):
    try: await c.answer()
    except: pass

    uid = c.from_user.id
    correct, total = await get_stats(uid)
    u = await get_user(uid)
    pct = round(correct / total * 100) if total else 0
    bar = "🟩" * (pct // 10) + "⬜" * (10 - pct // 10)

    is_admin_user = uid in ADMIN_IDS
    menu_kb = admin_menu_kb() if is_admin_user else user_menu_kb()

    text = (
        f"📊 <b>Sizning statistikangiz</b>\n\n"
        f"👤 {u[2] if u and len(u) > 2 else '-'} {u[3] if u and len(u) > 3 else ''}\n"
        f"📚 Daraja: {u[4] if u and len(u) > 4 else '-'}\n\n"
        f"✅ To'g'ri: <b>{correct}</b>\n"
        f"❌ Noto'g'ri: <b>{total - correct}</b>\n"
        f"📝 Jami: <b>{total}</b>\n"
        f"🎯 Natija: <b>{pct}%</b>\n\n"
        f"{bar}"
    )
    try: await c.message.edit_text(text, reply_markup=menu_kb)
    except TelegramBadRequest: await c.message.answer(text, reply_markup=menu_kb)

# ═══════════════════════════════════════════════════════════════
# 👨‍💼  ADMIN PANEL
# ═══════════════════════════════════════════════════════════════
def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS

@dp.callback_query(F.data == "admin_panel")
async def cb_admin_panel(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if not is_admin(c.from_user.id):
        return await c.message.answer("🚫 Sizda ruxsat yo'q!")

    users = await all_users()
    text = f"👨‍💼 <b>Admin Panel</b>\n\n👥 Jami foydalanuvchilar: <b>{len(users)}</b>\n\nNima olmoqchisiz? 👇"
    try: await c.message.edit_text(text, reply_markup=admin_kb())
    except TelegramBadRequest: await c.message.answer(text, reply_markup=admin_kb())

@dp.callback_query(F.data == "admin_excel")
async def cb_admin_excel(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if not is_admin(c.from_user.id):
        return await c.message.answer("🚫 Ruxsat yo'q!")

    await c.message.answer("⏳ Excel tayyorlanmoqda...")
    users = await all_users()
    data = await asyncio.to_thread(build_excel, users)  # ⚡ Thread ichida
    fname = f"statistika_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    await bot.send_document(
        c.from_user.id,
        BufferedInputFile(data, filename=fname),
        caption=f"📊 Statistika — {len(users)} foydalanuvchi"
    )

@dp.callback_query(F.data == "admin_txt")
async def cb_admin_txt(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if not is_admin(c.from_user.id):
        return await c.message.answer("🚫 Ruxsat yo'q!")

    await c.message.answer("⏳ Text fayl tayyorlanmoqda...")
    users = await all_users()
    text = build_txt(users)
    fname = f"statistika_{datetime.now().strftime('%d%m%Y_%H%M')}.txt"
    await bot.send_document(
        c.from_user.id,
        BufferedInputFile(text.encode("utf-8"), filename=fname),
        caption=f"📝 Statistika — {len(users)} foydalanuvchi"
    )

@dp.callback_query(F.data == "admin_count")
async def cb_admin_count(c: CallbackQuery):
    try: await c.answer()
    except: pass
    if not is_admin(c.from_user.id):
        return await c.message.answer("🚫 Ruxsat yo'q!")

    users = await all_users()
    reg = sum(1 for u in users if u and len(u) > 6 and u[6] and u[6] > 0)
    await c.message.answer(f"👥 Jami: {len(users)}\n✅ Faol: {reg}")

@dp.callback_query(F.data == "admin_broadcast")
async def cb_admin_broadcast(c: CallbackQuery, state: FSMContext):
    try: await c.answer()
    except: pass
    if not is_admin(c.from_user.id):
        return await c.message.answer("🚫 Ruxsat yo'q!")

    await state.set_state(S.broadcast)
    await c.message.answer("📢 Yuboriladigan xabarni kiriting:\n(Matn, rasm yoki hujjat)")

@dp.message(S.broadcast)
async def do_broadcast(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    users = await all_users()
    ok, fail = 0, 0
    status = await message.answer(f"📢 {len(users)} ta foydalanuvchiga yuborilmoqda...")
    
    for u in users:
        uid = u[0]
        try:
            await message.copy_to(uid)
            ok += 1
            await asyncio.sleep(0.03)  # ⚡ Telegram limitdan qochish
        except Exception:
            fail += 1
    
    await status.edit_text(
        f"✅ Broadcast yakunlandi!\n"
        f"📨 Muvaffaqiyatli: {ok}\n❌ Xato: {fail}"
    )

# ═══════════════════════════════════════════════════════════════
# 🎯  JAVOB HANDLER
# ═══════════════════════════════════════════════════════════════
@dp.callback_query(F.data.startswith("ans_"))
async def handle_answer(c: CallbackQuery):
    try: await c.answer()
    except: pass

    uid = c.from_user.id
    if uid not in sessions:
        return await c.message.answer("⚠️ Session topilmadi. /start bosing.")

    d = sessions[uid]
    q = d["tests"][d["index"]]
    sel = int(c.data.split("_")[1])
    practice = d.get("practice", False)

    if sel == q["correct"]:
        d["score"] += 1
        result_line = "\n\n✅ <b>To'g'ri!</b>"
        if not practice:
            await update_stats(uid, 1)
    else:
        ct = q["options"][q["correct"]]
        result_line = f"\n\n❌ <b>Noto'g'ri!</b>\n💡 To'g'ri: <b>{ct}</b>"
        if not practice:
            await update_stats(uid, 0)

    try:
        await c.message.edit_text(c.message.text + result_line)
    except TelegramBadRequest:
        pass

    d["index"] += 1
    if d["index"] < len(d["tests"]):
        await send_question(c.message.chat.id, uid)
    else:
        score = d["score"]
        total = len(d["tests"])
        pct = round(score / total * 100)

        if pct == 100:  medal = "🏆 Mukammal!"
        elif pct >= 80: medal = "🥇 A'lo!"
        elif pct >= 60: medal = "🥈 Yaxshi!"
        elif pct >= 40: medal = "🥉 O'rtacha"
        else:           medal = "📚 Ko'proq o'qing!"

        test_id = d.get("test_id")
        test_title = d.get("title", "Test")

        if test_id and not practice:
            await save_result(test_id, uid, score, total)
            # ⚡ Admin notification ORQADA ishlaydi — bot qotmaydi
            asyncio.create_task(notify_admins(uid, test_title, score, total))

        is_admin_user = uid in ADMIN_IDS
        menu_kb = admin_menu_kb() if is_admin_user else user_menu_kb()

        practice_note = (
            "\n\n⚠️ <i>Bu — tayyorlanish urinishi edi, ball reytingga hisoblanmadi.</i>"
            if practice else ""
        )

        await c.message.answer(
            f"🏁 <b>{d['title']} — Yakuniy natija</b>\n\n"
            f"✅ To'g'ri: <b>{score}</b>\n"
            f"❌ Noto'g'ri: <b>{total - score}</b>\n"
            f"📊 Natija: <b>{score}/{total}</b> ({pct}%)\n\n"
            f"{medal}{practice_note}",
            reply_markup=menu_kb
        )
        del sessions[uid]

# ═══════════════════════════════════════════════════════════════
# 💬  GURUH XABARLARI
# ═══════════════════════════════════════════════════════════════
@dp.message(F.new_chat_members)
async def bot_added_to_group(message: Message):
    for member in message.new_chat_members:
        if member.id == bot.id:
            await save_group(message.chat.id, message.chat.title or "Guruh")
            await message.answer(
                "👋 <b>Salom, guruh a'zolari!</b>\n\n"
                "Men — Quiz Bot. 🤖\n"
                "Admin testlarni men orqali shu guruhga yuborishi mumkin.\n\n"
                "⚠️ Iltimos, menga <b>admin</b> huquqi bering!"
            )

# ═══════════════════════════════════════════════════════════════
# 📋  KOMANDALAR
# ═══════════════════════════════════════════════════════════════
@dp.message(Command("help"))
async def cmd_help(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "📖 <b>Yordam</b>\n\n"
        "/start — Asosiy menyu\n"
        "/stat  — Statistikam\n"
        "/help  — Yordam\n"
        "/dev   — Dasturchi\n\n"
        "<b>Fayl turlari:</b> PDF, DOCX, TXT, JPG, PNG\n\n"
        "<b>Matn formati:</b>\n"
        "<code>1. Savol?\nA) ...\nB) ...\nC) ...\nD) ...</code>"
    )

@dp.message(Command("stat"))
async def cmd_stat(message: Message):
    uid = message.from_user.id
    correct, total = await get_stats(uid)
    u = await get_user(uid)
    pct = round(correct / total * 100) if total else 0
    bar = "🟩" * (pct // 10) + "⬜" * (10 - pct // 10)
    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👤 {u[2] if u and len(u) > 2 else '-'} {u[3] if u and len(u) > 3 else ''}\n"
        f"✅ To'g'ri: {correct}\n❌ Noto'g'ri: {total-correct}\n"
        f"📝 Jami: {total}\n🎯 {pct}%\n{bar}"
    )

@dp.message(Command("dev"))
async def cmd_dev(message: Message):
    b = InlineKeyboardBuilder()
    b.button(text="👨‍💻 Dasturchi", url=f"https://t.me/{ADMIN_USERNAME}")
    await message.answer(
        "🛠 <b>Yaratuvchi:</b> Sultonboyev Muhammad\n"
        f"📬 @{ADMIN_USERNAME}",
        reply_markup=b.as_markup()
    )

@dp.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        return await message.answer("🚫 Ruxsat yo'q!")
    users = await all_users()
    await message.answer(
        f"👨‍💼 <b>Admin Panel</b>\n👥 Jami: <b>{len(users)}</b> foydalanuvchi",
        reply_markup=admin_kb()
    )

# ═══════════════════════════════════════════════════════════════
# 🔄  FALLBACK
# ═══════════════════════════════════════════════════════════════
@dp.message(F.document | F.photo)
async def fallback_file(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.set_state(S.wait_file)
    await handle_file(message, state)

@dp.message(F.text)
async def fallback_text(message: Message):
    if message.text.startswith("/"):
        return
    if message.chat.type in ("group", "supergroup"):
        return
    await message.answer(
        "Iltimos, menyu orqali tanlang 👇",
        reply_markup=admin_menu_kb() if message.from_user.id in ADMIN_IDS else user_menu_kb()
    )

# ═══════════════════════════════════════════════════════════════
# 🛡️  GLOBAL ERROR HANDLER
# ═══════════════════════════════════════════════════════════════
@dp.error()
async def on_error(update, exception):
    log.error(f"Global error: {exception}", exc_info=exception)
    return True

# ═══════════════════════════════════════════════════════════════
# ▶️  ISHGA TUSHIRISH
# ══════════════════════════════════════════════════════════════
async def on_startup():
    await init_db()
    log.info("=" * 50)
    log.info("✅ Quiz Bot Professional (OPTIMIZED) ishga tushdi!")
    log.info("=" * 50)

async def on_shutdown():
    await close_resources()
    log.info("Bot to'xtatildi. Resurslar tozalandi.")

async def main():
    dp.startup.register(on_startup)
    dp.shutdown.register(on_shutdown)
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())
